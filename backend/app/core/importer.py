"""
========================================
通用 Excel/CSV 数据导入框架
========================================

为所有模块提供统一的模板生成、文件解析、数据校验和批量导入功能。

设计原则：
1. 简洁：每个实体只保留最核心字段
2. 安全：严格的类型校验和必填检查
3. 容错：中文逗号兼容、多余空格去除、常见格式容错
4. 透明：详细的逐行错误报告

使用方法：
    1. 定义 ImportField 列表描述字段规则
    2. 调用 BaseImporter 的方法生成模板/解析/导入
    3. 在模块 service_import.py 中封装具体业务逻辑
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any, Callable, Dict, List, Optional, Tuple, Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session


# ── 数据模型 ───────────────────────────────────────────


@dataclass
class ImportField:
    """导入字段定义"""
    key: str                          # 数据库字段名（英文 snake_case）
    header: str                       # Excel 表头（中文）
    required: bool = False            # 是否必填
    field_type: str = "str"           # str | int | float | bool | enum | list | date
    enum_values: Optional[List[str]] = None   # 枚举值列表（英文值）
    enum_display: Optional[List[str]] = None  # 枚举值显示（中文，可选）
    default: Any = None               # 默认值
    description: str = ""             # 字段说明（用于模板注释）
    example: Any = ""                 # 示例值
    max_length: Optional[int] = None  # 最大长度
    min_value: Optional[Union[int, float]] = None
    max_value: Optional[Union[int, float]] = None
    allow_empty_as_none: bool = True  # 空字符串是否转为 None
    custom_validator: Optional[Callable[[Any], Tuple[bool, str]]] = None


@dataclass
class ImportRow:
    """解析后的单行数据"""
    row_number: int                   # Excel/CSV 中的行号（从 2 开始）
    action: str                       # "IMPORT" | "SKIP"
    data: Dict[str, Any]              # 解析后的字段数据（空值已剔除）
    raw_identifier: Optional[str] = None  # 用于错误报告的标识（如姓名、学号）


@dataclass
class ImportErrorItem:
    """导入错误项"""
    row_number: int
    identifier: Optional[str]
    message: str


@dataclass
class ImportResult:
    """导入结果统计"""
    created: int = 0
    updated: int = 0
    skipped: int = 0
    failed: int = 0
    errors: List[ImportErrorItem] = field(default_factory=list)

    def ok(self) -> bool:
        return self.failed == 0


# ── 基础工具函数 ───────────────────────────────────────


def _norm_str(v: Any) -> str:
    """标准化字符串值"""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _parse_action(v: Any) -> str:
    """解析导入标记"""
    s = _norm_str(v).upper()
    if not s:
        return "IMPORT"
    if s in {"IMPORT", "SKIP", "CREATE", "UPDATE"}:
        return "IMPORT"  # CREATE/UPDATE 统一为 IMPORT
    if s in {"导入", "导入更新", "更新", "创建"}:
        return "IMPORT"
    if s in {"跳过", "忽略"}:
        return "SKIP"
    return "INVALID"


def _parse_str(v: Any, max_length: Optional[int] = None) -> Tuple[Optional[str], Optional[str]]:
    """解析字符串，返回 (值, 错误信息)"""
    s = _norm_str(v)
    if not s:
        return (None, None)
    if max_length is not None and len(s) > max_length:
        return (None, f"长度超过限制（最大 {max_length} 字符）")
    return (s, None)


def _parse_int(v: Any, min_v: Optional[int] = None, max_v: Optional[int] = None) -> Tuple[Optional[int], Optional[str]]:
    """解析整数，返回 (值, 错误信息)"""
    s = _norm_str(v)
    if not s:
        return (None, None)
    try:
        # Excel 中数字可能是 float，如 25.0
        val = int(float(s))
    except (ValueError, TypeError):
        return (None, "必须为整数")
    if min_v is not None and val < min_v:
        return (None, f"不能小于 {min_v}")
    if max_v is not None and val > max_v:
        return (None, f"不能大于 {max_v}")
    return (val, None)


def _parse_float(v: Any, min_v: Optional[float] = None, max_v: Optional[float] = None) -> Tuple[Optional[float], Optional[str]]:
    """解析浮点数，返回 (值, 错误信息)"""
    s = _norm_str(v)
    if not s:
        return (None, None)
    try:
        val = float(s)
    except (ValueError, TypeError):
        return (None, "必须为数字")
    if min_v is not None and val < min_v:
        return (None, f"不能小于 {min_v}")
    if max_v is not None and val > max_v:
        return (None, f"不能大于 {max_v}")
    return (val, None)


def _parse_bool(v: Any) -> Tuple[Optional[bool], Optional[str]]:
    """解析布尔值，返回 (值, 错误信息)"""
    s = _norm_str(v)
    if not s:
        return (None, None)
    sl = s.lower()
    if sl in {"true", "1", "yes", "是", "y", "t", "真", "对"}:
        return (True, None)
    if sl in {"false", "0", "no", "否", "n", "f", "假", "错", "", " "}:
        return (False, None)
    return (None, "布尔值格式错误，请填写：是/否、1/0、true/false")


def _parse_enum(v: Any, enum_values: List[str], enum_display: Optional[List[str]] = None) -> Tuple[Optional[str], Optional[str]]:
    """解析枚举值，支持通过中文显示值匹配，返回 (英文值, 错误信息)"""
    s = _norm_str(v)
    if not s:
        return (None, None)
    # 直接匹配英文值
    if s in enum_values:
        return (s, None)
    # 通过中文显示值匹配
    if enum_display:
        for ev, ed in zip(enum_values, enum_display):
            if s == ed:
                return (ev, None)
    # 大小写不敏感匹配
    s_upper = s.upper()
    for ev in enum_values:
        if s_upper == ev.upper():
            return (ev, None)
    allowed = ", ".join(enum_values)
    if enum_display:
        pairs = [f"{ev}({ed})" for ev, ed in zip(enum_values, enum_display)]
        allowed = ", ".join(pairs)
    return (None, f"无效值，可选：{allowed}")


def _parse_list(v: Any) -> Tuple[Optional[List[str]], Optional[str]]:
    """解析列表（逗号分隔），返回 (值, 错误信息)"""
    s = _norm_str(v)
    if not s:
        return (None, None)
    # 兼容中英文逗号
    parts = []
    for token in s.replace("，", ",").split(","):
        t = token.strip()
        if t:
            parts.append(t)
    return (parts if parts else None, None)


def _parse_date(v: Any) -> Tuple[Optional[Union[date, datetime]], Optional[str]]:
    """解析日期，返回 (值, 错误信息)"""
    s = _norm_str(v)
    if not s:
        return (None, None)
    # Excel 日期可能是 datetime 对象
    if isinstance(v, (date, datetime)):
        return (v, None)
    # 尝试多种格式
    formats = ["%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%d-%m-%Y", "%d/%m/%Y"]
    for fmt in formats:
        try:
            return (datetime.strptime(s, fmt).date(), None)
        except ValueError:
            continue
    return (None, "日期格式错误，请使用 YYYY-MM-DD 格式")


# ── 核心导入器 ─────────────────────────────────────────


class BaseImporter:
    """
    通用数据导入器

    每个模块创建实例时传入字段配置，即可获得：
    - 模板生成（xlsx/csv）
    - 文件解析
    - 数据校验
    """

    def __init__(
        self,
        fields: List[ImportField],
        sheet_name: str = "导入模板",
        id_field_key: Optional[str] = None,  # 唯一标识字段（如 code, student_no, name）
    ):
        self.fields = [f for f in fields if f.header]  # 过滤掉无表头的字段
        self.sheet_name = sheet_name
        self.id_field_key = id_field_key
        self.headers = [f.header for f in self.fields]

    # ── 模板生成 ──────────────────────────────────────

    def build_template_xlsx(self) -> bytes:
        """生成 Excel 导入模板（含表头、示例、说明）"""
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name

        # 样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        desc_font = Font(italic=True, color="666666")
        desc_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # 第 1 行：表头
        for col_idx, field_def in enumerate(self.fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=field_def.header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 第 2 行：说明（描述 + 约束）
        for col_idx, field_def in enumerate(self.fields, 1):
            parts = []
            if field_def.description:
                parts.append(field_def.description)
            if field_def.required:
                parts.append("【必填】")
            if field_def.field_type == "enum" and field_def.enum_values:
                if field_def.enum_display:
                    pairs = [f"{ev}({ed})" for ev, ed in zip(field_def.enum_values, field_def.enum_display)]
                    parts.append(f"可选：{', '.join(pairs)}")
                else:
                    parts.append(f"可选：{', '.join(field_def.enum_values)}")
            if field_def.max_length:
                parts.append(f"最长{field_def.max_length}字符")
            if field_def.min_value is not None or field_def.max_value is not None:
                range_str = ""
                if field_def.min_value is not None:
                    range_str += f"≥{field_def.min_value}"
                if field_def.max_value is not None:
                    range_str += f"≤{field_def.max_value}"
                parts.append(f"范围：{range_str}")

            desc = " | ".join(parts)
            cell = ws.cell(row=2, column=col_idx, value=desc)
            cell.font = desc_font
            cell.fill = desc_fill
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = thin_border

        # 第 3 行：示例
        for col_idx, field_def in enumerate(self.fields, 1):
            example = field_def.example if field_def.example is not None else ""
            cell = ws.cell(row=3, column=col_idx, value=example)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # 第 4 行：示例 2（不同值）
        for col_idx, field_def in enumerate(self.fields, 1):
            # 根据字段类型生成第二个示例
            example2 = self._generate_second_example(field_def)
            cell = ws.cell(row=4, column=col_idx, value=example2)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # 自适应列宽
        for col_idx, field_def in enumerate(self.fields, 1):
            col_letter = get_column_letter(col_idx)
            # 基于表头、说明、示例中最长内容
            max_width = max(
                len(str(field_def.header)),
                len(str(field_def.description or "")),
                len(str(field_def.example or "")),
                10
            )
            ws.column_dimensions[col_letter].width = min(max_width + 4, 40)

        # 行高
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 45

        # 冻结首行
        ws.freeze_panes = "A2"

        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()

    def _generate_second_example(self, field_def: ImportField) -> str:
        """生成第二行示例数据"""
        if field_def.example is not None:
            ex = str(field_def.example)
            # 尝试生成不同的示例
            if field_def.field_type == "enum" and field_def.enum_values:
                if ex in field_def.enum_values and len(field_def.enum_values) > 1:
                    for ev in field_def.enum_values:
                        if ev != ex:
                            return ev
            if field_def.field_type == "bool":
                return "否" if ex in ("是", True, "1", "true") else "是"
            if field_def.field_type == "int" and isinstance(field_def.example, int):
                return str(field_def.example + 1)
        return ""

    def build_template_csv(self) -> bytes:
        """生成 CSV 导入模板"""
        sio = StringIO()
        writer = csv.writer(sio)
        writer.writerow([f.header for f in self.fields])
        writer.writerow([f.example if f.example is not None else "" for f in self.fields])
        return ("\ufeff" + sio.getvalue()).encode("utf-8")

    # ── 文件解析 ──────────────────────────────────────

    def parse_file(self, filename: str, content: bytes) -> Tuple[List[ImportRow], List[ImportErrorItem]]:
        """解析上传文件，返回 (有效行, 错误列表)"""
        lower = filename.lower()
        if lower.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            return self._parse_xlsx(content)
        if lower.endswith(".csv"):
            return self._parse_csv(content)
        return ([], [ImportErrorItem(row_number=0, identifier=None, message="仅支持 .xlsx 或 .csv 文件")])

    def _parse_xlsx(self, content: bytes) -> Tuple[List[ImportRow], List[ImportErrorItem]]:
        """解析 Excel 文件"""
        errors: List[ImportErrorItem] = []
        rows: List[ImportRow] = []

        try:
            wb = load_workbook(filename=BytesIO(content), data_only=True)
        except Exception as e:
            return ([], [ImportErrorItem(row_number=0, identifier=None, message=f"Excel 文件解析失败: {str(e)}")])

        ws = wb.active

        # 读取表头（第 1 行）
        first_row = list(ws.iter_rows(min_row=1, max_row=1))
        if not first_row:
            return ([], [ImportErrorItem(row_number=1, identifier=None, message="文件为空，缺少表头")])

        header_row = [str(c.value).strip() if c.value is not None else "" for c in first_row[0]]
        header_index: Dict[str, int] = {}
        for idx, h in enumerate(header_row):
            if h:
                header_index[h] = idx

        # 检查必填表头
        missing = [f.header for f in self.fields if f.header not in header_index]
        if missing:
            return ([], [ImportErrorItem(row_number=1, identifier=None, message=f"缺少表头列：{', '.join(missing)}")])

        # 从第 2 行开始读取（跳过第 2 行的说明和第 3-4 行的示例，但保留容错）
        for excel_row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # 跳过完全空行
            all_none = all(c.value is None or str(c.value).strip() == "" for c in row)
            if all_none:
                continue

            result = self._parse_row(row, header_index, excel_row_idx)
            if result[0]:  # ImportRow
                rows.append(result[0])
            if result[1]:  # ImportErrorItem
                errors.append(result[1])

        return (rows, errors)

    def _parse_csv(self, content: bytes) -> Tuple[List[ImportRow], List[ImportErrorItem]]:
        """解析 CSV 文件"""
        errors: List[ImportErrorItem] = []
        rows: List[ImportRow] = []

        try:
            text = content.decode("utf-8-sig", errors="replace")
        except Exception as e:
            return ([], [ImportErrorItem(row_number=0, identifier=None, message=f"文件编码错误: {str(e)}")])

        reader = csv.DictReader(StringIO(text))

        if reader.fieldnames is None:
            return ([], [ImportErrorItem(row_number=1, identifier=None, message="CSV 缺少表头")])

        missing = [f.header for f in self.fields if f.header not in reader.fieldnames]
        if missing:
            return ([], [ImportErrorItem(row_number=1, identifier=None, message=f"缺少表头列：{', '.join(missing)}")])

        csv_row_number = 1
        for record in reader:
            csv_row_number += 1
            # 跳过完全空行
            if all(v is None or str(v).strip() == "" for v in record.values()):
                continue

            result = self._parse_record(record, csv_row_number)
            if result[0]:
                rows.append(result[0])
            if result[1]:
                errors.append(result[1])

        return (rows, errors)

    def _parse_row(
        self, row: tuple, header_index: Dict[str, int], row_number: int
    ) -> Tuple[Optional[ImportRow], Optional[ImportErrorItem]]:
        """解析 Excel 单行"""
        data: Dict[str, Any] = {}
        identifier: Optional[str] = None

        # 提取标识字段值（用于错误报告）
        if self.id_field_key:
            id_field = next((f for f in self.fields if f.key == self.id_field_key), None)
            if id_field and id_field.header in header_index:
                identifier = _norm_str(row[header_index[id_field.header]].value)

        # 解析导入标记（如果有的话）
        action = "IMPORT"
        action_header = next((f.header for f in self.fields if f.key == "action"), None)
        if action_header and action_header in header_index:
            action = _parse_action(row[header_index[action_header]].value)
            if action == "INVALID":
                return (None, ImportErrorItem(
                    row_number=row_number, identifier=identifier or "",
                    message="导入标记无效，请填写 IMPORT 或 SKIP（可留空）"
                ))

        if action == "SKIP":
            return (ImportRow(row_number=row_number, action="SKIP", data={}, raw_identifier=identifier), None)

        # 逐字段解析
        for field_def in self.fields:
            if field_def.header not in header_index:
                continue
            cell_value = row[header_index[field_def.header]].value
            parsed, error = self._parse_field_value(cell_value, field_def)
            if error:
                return (None, ImportErrorItem(
                    row_number=row_number,
                    identifier=identifier or "",
                    message=f"【{field_def.header}】{error}"
                ))
            if parsed is not None:
                data[field_def.key] = parsed

        return (ImportRow(row_number=row_number, action="IMPORT", data=data, raw_identifier=identifier), None)

    def _parse_record(
        self, record: Dict[str, str], row_number: int
    ) -> Tuple[Optional[ImportRow], Optional[ImportErrorItem]]:
        """解析 CSV 单行"""
        data: Dict[str, Any] = {}
        identifier: Optional[str] = None

        # 提取标识字段值
        if self.id_field_key:
            id_field = next((f for f in self.fields if f.key == self.id_field_key), None)
            if id_field:
                identifier = _norm_str(record.get(id_field.header))

        # 解析导入标记
        action = "IMPORT"
        action_header = next((f.header for f in self.fields if f.key == "action"), None)
        if action_header:
            action = _parse_action(record.get(action_header))
            if action == "INVALID":
                return (None, ImportErrorItem(
                    row_number=row_number, identifier=identifier or "",
                    message="导入标记无效，请填写 IMPORT 或 SKIP（可留空）"
                ))

        if action == "SKIP":
            return (ImportRow(row_number=row_number, action="SKIP", data={}, raw_identifier=identifier), None)

        # 逐字段解析
        for field_def in self.fields:
            raw_value = record.get(field_def.header)
            parsed, error = self._parse_field_value(raw_value, field_def)
            if error:
                return (None, ImportErrorItem(
                    row_number=row_number,
                    identifier=identifier or "",
                    message=f"【{field_def.header}】{error}"
                ))
            if parsed is not None:
                data[field_def.key] = parsed

        return (ImportRow(row_number=row_number, action="IMPORT", data=data, raw_identifier=identifier), None)

    def _parse_field_value(self, raw_value: Any, field_def: ImportField) -> Tuple[Any, Optional[str]]:
        """解析单个字段值，返回 (解析后的值, 错误信息)"""
        s = _norm_str(raw_value)

        # 空值处理
        if not s:
            if field_def.required:
                return (None, "必填项不能为空")
            return (field_def.default, None) if field_def.default is not None else (None, None)

        # 自定义校验器
        if field_def.custom_validator:
            ok, msg = field_def.custom_validator(raw_value)
            if not ok:
                return (None, msg)

        # 类型解析
        if field_def.field_type == "str":
            val, err = _parse_str(raw_value, field_def.max_length)
            return (val, err)

        elif field_def.field_type == "int":
            val, err = _parse_int(raw_value, field_def.min_value, field_def.max_value)
            return (val, err)

        elif field_def.field_type == "float":
            val, err = _parse_float(raw_value, field_def.min_value, field_def.max_value)
            return (val, err)

        elif field_def.field_type == "bool":
            val, err = _parse_bool(raw_value)
            return (val, err)

        elif field_def.field_type == "enum":
            if not field_def.enum_values:
                return (s, None)
            val, err = _parse_enum(raw_value, field_def.enum_values, field_def.enum_display)
            return (val, err)

        elif field_def.field_type == "list":
            val, err = _parse_list(raw_value)
            return (val, err)

        elif field_def.field_type == "date":
            val, err = _parse_date(raw_value)
            return (val, err)

        # 默认字符串
        val, err = _parse_str(raw_value, field_def.max_length)
        return (val, err)

    # ── 重复检查 ──────────────────────────────────────

    def validate_unique_in_file(
        self, rows: List[ImportRow], unique_key: str
    ) -> List[ImportErrorItem]:
        """检查文件内是否有重复的唯一键值"""
        errors: List[ImportErrorItem] = []
        seen: Dict[Any, int] = {}
        for r in rows:
            if r.action == "SKIP":
                continue
            val = r.data.get(unique_key)
            if val is None:
                continue
            if val in seen:
                errors.append(ImportErrorItem(
                    row_number=r.row_number,
                    identifier=str(val),
                    message=f"文件内存在重复：第 {seen[val]} 行和第 {r.row_number} 行的 '{unique_key}' 值相同"
                ))
            else:
                seen[val] = r.row_number
        return errors


# ── 快捷函数 ───────────────────────────────────────────


def create_import_response(result: ImportResult) -> dict:
    """将 ImportResult 转为标准 API 响应"""
    return {
        "code": 200 if result.ok() else 200,
        "message": "导入完成" if result.ok() else "导入完成（部分失败）",
        "data": {
            "created": result.created,
            "updated": result.updated,
            "skipped": result.skipped,
            "failed": result.failed,
            "errors": [{"rowNumber": e.row_number, "identifier": e.identifier, "message": e.message} for e in result.errors],
        },
    }
