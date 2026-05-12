"""
教师模块业务逻辑

包含 Excel/CSV 模板生成与批量导入解析逻辑。
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import Any, Dict, Iterable, List, Optional, Tuple

import csv

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.modules.teachers import crud
from app.modules.teachers.models import ResearchGroup
from app.modules.teachers.schemas import TeacherCreate, TeacherUpdate


TEMPLATE_HEADERS: List[str] = [
    "姓名",
    "导入标记",
    "教师类型",
    "学部",
    "每周最大课时",
    "标签",
    "教研组",
]


def build_teachers_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "教师导入模板"

    ws.append(TEMPLATE_HEADERS)
    ws.append(
        [
            "张三",
            "IMPORT",
            "CN",
            "PRIMARY",
            25,
            "PRIMARY_ADMIN",
            "数学教研组",
        ]
    )
    ws.append(
        [
            "李四",
            "",
            "EN",
            "SECONDARY",
            18,
            "ASSISTANT_HOMEROOM,SECONDARY_ADMIN",
            "英语教研组",
        ]
    )
    ws.append(["王五", "SKIP", "", "", "", "", ""])

    # 简单自适应列宽
    for col_idx in range(1, len(TEMPLATE_HEADERS) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 16

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_teachers_template_csv() -> bytes:
    sio = StringIO()
    writer = csv.writer(sio)
    writer.writerow(TEMPLATE_HEADERS)
    writer.writerow(
        [
            "张三",
            "IMPORT",
            "CN",
            "PRIMARY",
            "25",
            "PRIMARY_ADMIN",
            "数学教研组",
        ]
    )
    writer.writerow(
        [
            "李四",
            "",
            "EN",
            "SECONDARY",
            "18",
            "ASSISTANT_HOMEROOM,SECONDARY_ADMIN",
            "英语教研组",
        ]
    )
    writer.writerow(["王五", "SKIP", "", "", "", "", ""])
    # UTF-8 with BOM for Excel-friendly CSV
    return ("\ufeff" + sio.getvalue()).encode("utf-8")


@dataclass(frozen=True)
class ImportRow:
    row_number: int
    name: str
    action: str  # IMPORT or SKIP
    payload: Dict[str, Any]  # fields for TeacherCreate/TeacherUpdate (no empty values)


@dataclass(frozen=True)
class ImportErrorItem:
    row_number: int
    name: Optional[str]
    message: str


@dataclass(frozen=True)
class ImportResult:
    created: int
    updated: int
    skipped: int
    failed: int
    errors: List[ImportErrorItem]


def _norm_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _parse_action(v: Any) -> str:
    s = _norm_str(v).upper()
    if not s:
        return "IMPORT"
    if s in {"IMPORT", "SKIP"}:
        return s
    # 容错：允许中文
    if s in {"导入", "导入更新", "更新"}:
        return "IMPORT"
    if s in {"跳过", "忽略"}:
        return "SKIP"
    return "INVALID"


def _split_tags(v: Any) -> Optional[List[str]]:
    s = _norm_str(v)
    if not s:
        return None
    parts = []
    for token in s.replace("，", ",").split(","):
        t = token.strip()
        if t:
            parts.append(t)
    if not parts:
        return None
    # 按你的要求：禁止/忽略班主任标签
    parts = [t for t in parts if t != "HOMEROOM_TEACHER"]
    return parts or None


def _parse_int_optional(v: Any) -> Optional[int]:
    s = _norm_str(v)
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _get_or_create_research_group_id(db: Session, group_name: str) -> int:
    name = group_name.strip()
    group = (
        db.query(ResearchGroup)
        .filter(ResearchGroup.is_deleted == False, ResearchGroup.name == name)
        .first()
    )
    if group:
        return int(group.id)
    group = ResearchGroup(name=name)
    db.add(group)
    db.commit()
    db.refresh(group)
    return int(group.id)


def parse_teachers_import_file(filename: str, content: bytes) -> Tuple[List[ImportRow], List[ImportErrorItem]]:
    lower = filename.lower()
    if lower.endswith(".xlsx") or lower.endswith(".xlsm") or lower.endswith(".xltx") or lower.endswith(".xltm"):
        return _parse_xlsx(content)
    if lower.endswith(".csv"):
        return _parse_csv(content)
    return ([], [ImportErrorItem(row_number=0, name=None, message="仅支持 .xlsx 或 .csv 文件")])


def _parse_xlsx(content: bytes) -> Tuple[List[ImportRow], List[ImportErrorItem]]:
    errors: List[ImportErrorItem] = []
    rows: List[ImportRow] = []

    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active

    header_row = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_index = {h: idx for idx, h in enumerate(header_row)}

    missing = [h for h in TEMPLATE_HEADERS if h not in header_index]
    if missing:
        return ([], [ImportErrorItem(row_number=1, name=None, message=f"缺少表头列：{', '.join(missing)}")])

    for excel_row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        name = _norm_str(row[header_index["姓名"]].value)
        if not name:
            errors.append(ImportErrorItem(row_number=excel_row_idx, name=None, message="姓名为空"))
            continue

        action = _parse_action(row[header_index["导入标记"]].value)
        if action == "INVALID":
            errors.append(ImportErrorItem(row_number=excel_row_idx, name=name, message="导入标记仅支持 IMPORT/SKIP（可留空）"))
            continue

        if action == "SKIP":
            rows.append(ImportRow(row_number=excel_row_idx, name=name, action=action, payload={}))
            continue

        payload: Dict[str, Any] = {}

        teacher_type = _norm_str(row[header_index["教师类型"]].value)
        if teacher_type:
            payload["type"] = teacher_type

        department = _norm_str(row[header_index["学部"]].value)
        if department:
            payload["department"] = department

        max_weekly_hours = _parse_int_optional(row[header_index["每周最大课时"]].value)
        if max_weekly_hours is None and _norm_str(row[header_index["每周最大课时"]].value):
            errors.append(ImportErrorItem(row_number=excel_row_idx, name=name, message="每周最大课时不是数字"))
            continue
        if max_weekly_hours is not None:
            payload["max_weekly_hours"] = max_weekly_hours

        tags = _split_tags(row[header_index["标签"]].value)
        if tags is not None:
            payload["tags"] = tags

        group_name = _norm_str(row[header_index["教研组"]].value)
        if group_name:
            payload["research_group_name"] = group_name

        rows.append(ImportRow(row_number=excel_row_idx, name=name, action="IMPORT", payload=payload))

    return (rows, errors)


def _parse_csv(content: bytes) -> Tuple[List[ImportRow], List[ImportErrorItem]]:
    errors: List[ImportErrorItem] = []
    rows: List[ImportRow] = []

    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(StringIO(text))

    if reader.fieldnames is None:
        return ([], [ImportErrorItem(row_number=1, name=None, message="CSV 缺少表头")])

    missing = [h for h in TEMPLATE_HEADERS if h not in reader.fieldnames]
    if missing:
        return ([], [ImportErrorItem(row_number=1, name=None, message=f"缺少表头列：{', '.join(missing)}")])

    csv_row_number = 1  # header is 1
    for record in reader:
        csv_row_number += 1
        name = _norm_str(record.get("姓名"))
        if not name:
            errors.append(ImportErrorItem(row_number=csv_row_number, name=None, message="姓名为空"))
            continue

        action = _parse_action(record.get("导入标记"))
        if action == "INVALID":
            errors.append(ImportErrorItem(row_number=csv_row_number, name=name, message="导入标记仅支持 IMPORT/SKIP（可留空）"))
            continue
        if action == "SKIP":
            rows.append(ImportRow(row_number=csv_row_number, name=name, action=action, payload={}))
            continue

        payload: Dict[str, Any] = {}

        teacher_type = _norm_str(record.get("教师类型"))
        if teacher_type:
            payload["type"] = teacher_type

        department = _norm_str(record.get("学部"))
        if department:
            payload["department"] = department

        max_weekly_hours_raw = record.get("每周最大课时")
        max_weekly_hours = _parse_int_optional(max_weekly_hours_raw)
        if max_weekly_hours is None and _norm_str(max_weekly_hours_raw):
            errors.append(ImportErrorItem(row_number=csv_row_number, name=name, message="每周最大课时不是数字"))
            continue
        if max_weekly_hours is not None:
            payload["max_weekly_hours"] = max_weekly_hours

        tags = _split_tags(record.get("标签"))
        if tags is not None:
            payload["tags"] = tags

        group_name = _norm_str(record.get("教研组"))
        if group_name:
            payload["research_group_name"] = group_name

        rows.append(ImportRow(row_number=csv_row_number, name=name, action="IMPORT", payload=payload))

    return (rows, errors)


def validate_duplicate_names(rows: List[ImportRow]) -> List[ImportErrorItem]:
    """
    重名规则（你确认的要求）：
    - 同名出现多行时：必须且只能有 1 行为 IMPORT，其余全部为 SKIP，否则阻止导入
    """
    errors: List[ImportErrorItem] = []
    by_name: Dict[str, List[ImportRow]] = {}
    for r in rows:
        by_name.setdefault(r.name, []).append(r)

    for name, group in by_name.items():
        if len(group) <= 1:
            continue
        import_rows = [r for r in group if r.action == "IMPORT"]
        skip_rows = [r for r in group if r.action == "SKIP"]
        if len(import_rows) == 1 and len(skip_rows) == (len(group) - 1):
            continue
        # 报错：列出涉及的行号，便于手工标注
        row_nums = ",".join(str(r.row_number) for r in group)
        errors.append(
            ImportErrorItem(
                row_number=import_rows[0].row_number if import_rows else group[0].row_number,
                name=name,
                message=f"同名教师在文件中出现多行（行 {row_nums}），请将要排除的行标记为 SKIP，只保留 1 行 IMPORT",
            )
        )
    return errors


def import_teachers_from_rows(db: Session, rows: List[ImportRow]) -> ImportResult:
    created = 0
    updated = 0
    skipped = 0
    errors: List[ImportErrorItem] = []

    for r in rows:
        if r.action == "SKIP":
            skipped += 1
            continue

        payload = dict(r.payload)
        group_name = payload.pop("research_group_name", None)
        if group_name:
            try:
                payload["research_group_id"] = _get_or_create_research_group_id(db, group_name)
            except Exception:
                errors.append(ImportErrorItem(row_number=r.row_number, name=r.name, message="教研组创建/查询失败"))
                continue

        # 查重：以 name 唯一
        existing = crud.get_teacher_by_name(db, r.name)

        if existing:
            # 空值不覆盖：payload 已在解析阶段剔除了空字段
            try:
                update_model = TeacherUpdate(**payload)
            except Exception as e:
                errors.append(ImportErrorItem(row_number=r.row_number, name=r.name, message=f"数据校验失败：{str(e)}"))
                continue

            updated_obj = crud.update_teacher(db, existing.id, update_model)
            if not updated_obj:
                errors.append(ImportErrorItem(row_number=r.row_number, name=r.name, message="更新失败：教师不存在"))
                continue
            updated += 1
        else:
            # 创建：允许大部分为空，使用默认值
            try:
                create_model = TeacherCreate(name=r.name, **payload)
            except Exception as e:
                errors.append(ImportErrorItem(row_number=r.row_number, name=r.name, message=f"数据校验失败：{str(e)}"))
                continue
            crud.create_teacher(db, create_model)
            created += 1

    return ImportResult(
        created=created,
        updated=updated,
        skipped=skipped,
        failed=len(errors),
        errors=errors,
    )

